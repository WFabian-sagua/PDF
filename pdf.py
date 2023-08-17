#Pdf to Excel
#Fabian Sagua
#datacountai@proton.me

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import os

#Ubicación y nombre del PDF
pdf_filename = 'nombre de tu archivo.pdf'

pdf_path = os.path.join(os.path.dirname(__file__), pdf_filename)

#Abrir PDF
with pdfplumber.open(pdf_path) as pdf:
    primera_pagina = pdf.pages[0]
    tablas = primera_pagina.extract_tables()

#Extraer los nombres de las columnas desde la primera fila de la tabla
nombres_columnas = tablas[0][0]

#Convertir la tabla en un dataframe en pandas
data_frame = pd.DataFrame(tablas[0][1:], columns=nombres_columnas)

#Crear un libro de trabajo en Excel
libro_excel = Workbook()
hoja_excel = libro_excel.active

#Para dar formato en negrita y bordes a los encabezados
encabezado_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
encabezado_font = Font(color='FFFFFF', bold=True)
encabezado_border = Border(bottom=Side(border_style='thin'))

for col, encabezado in enumerate(nombres_columnas, start=1):
    celda = hoja_excel.cell(row=1, column=col, value=encabezado)
    celda.font = encabezado_font
    celda.fill = encabezado_fill
    celda.border = encabezado_border

#Agrenado los datos del dataframe a la hoja Excel
for fila_datos in data_frame.itertuples(index=False):
    hoja_excel.append(fila_datos)

#Dar formato a todas las celdas
bordes = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
for fila in hoja_excel.iter_rows(min_row=1, max_row=hoja_excel.max_row,
                                 min_col=1, max_col=hoja_excel.max_column):
    for celda in fila:
        celda.border = bordes

#Guardar el archivo excel
archivo_excel = 'datos_extraidos.xlsx'
libro_excel.save(archivo_excel)

print(f'Datos exportados con éxito a {archivo_excel}')